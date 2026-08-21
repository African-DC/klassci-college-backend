"""Le journal des versements : ce qu'il totalise, ce qu'il nomme, ce qu'il sort.

Ces tests appellent les fonctions et regardent ce qu'elles produisent — le
journal composé, le HTML du document officiel, le classeur relu cellule par
cellule. Aucun n'inspecte le code source : un test qui lit du code verrouille
la forme d'une implémentation, pas la justesse d'un document, et c'est
exactement comme ça qu'un document faux se retrouve protégé par sa propre
suite de tests.
"""

from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.services.exports.payments_journal_xlsx import generate_payments_journal_xlsx
from app.services.payments_journal_service import build_journal
from app.services.pdf.payments_journal import render_payments_journal_html

ECOLE = {
    "school_name": "Lycée Moderne Saint-Augustin",
    "address": "Boulevard Latrille, Cocody",
    "phone": "+225 27 22 44 55 66",
    "email": "secretariat@saint-augustin.ci",
    "primary_color": "#1B4D3E",
    "accent_color": "#C1440E",
    "logo_url": None,
}


def _caissier(user_id: int, prenom: str, nom: str) -> SimpleNamespace:
    return SimpleNamespace(
        id=user_id,
        role="staff",
        email=f"{prenom.lower()}.{nom.lower()}@saint-augustin.ci",
        staff_profile=SimpleNamespace(first_name=prenom, last_name=nom),
        teacher_profile=None,
        student_profile=None,
        parent_profile=None,
    )


SOPHIE = _caissier(12, "Sophie", "Yao")
MARIAM = _caissier(18, "Mariam", "Diallo")


def _versement(
    payment_id: int,
    *,
    montant: str,
    method: str = "cash",
    status: str = "completed",
    caissier: SimpleNamespace | None = SOPHIE,
    categories: tuple[str, ...] = ("Inscription",),
    heure: int = 9,
) -> SimpleNamespace:
    allocations = [
        SimpleNamespace(
            enrollment_fee=SimpleNamespace(
                fee_variant=SimpleNamespace(category=SimpleNamespace(name=nom))
            )
        )
        for nom in categories
    ]
    return SimpleNamespace(
        id=payment_id,
        enrollment_id=payment_id,
        enrollment=SimpleNamespace(
            student=SimpleNamespace(
                first_name="Aminata",
                last_name="Traoré",
                enrollment_number="2025-6A-014",
                photo_url=None,
            )
        ),
        enrollment_fee=None,
        allocations=allocations,
        student_name_snapshot="Traoré Aminata",
        student_matricule_snapshot="2025-6A-014",
        created_at=datetime(2026, 9, 15, heure, 30),
        amount=Decimal(montant),
        method=method,
        status=status,
        reference=f"REC-{payment_id}",
        received_by=caissier.id if caissier else None,
        received_by_user=caissier,
    )


def _journal(versements: list[SimpleNamespace], **kwargs: object):
    defaults: dict = {
        "period_label": "Du 14/09/2026 au 16/09/2026",
        "filters_label": "",
        "scope_label": "Toutes les caisses",
        "school": ECOLE,
        "total_found": len(versements),
        "issued_at": datetime(2026, 9, 17, 9, 42),
    }
    defaults.update(kwargs)
    return build_journal(versements, **defaults)


# ---------------------------------------------------------------------------
# Qui a encaissé
# ---------------------------------------------------------------------------


def test_chaque_ligne_porte_le_nom_de_qui_a_encaisse() -> None:
    """La question posée à une ligne de caisse est « qui », pas « quel numéro »."""
    journal = _journal([_versement(1, montant="50000", caissier=SOPHIE)])
    assert journal.lines[0].cashier == "Sophie Yao"


def test_un_versement_sans_encaisseur_reste_nomme() -> None:
    """Une case vide au milieu d'une colonne de responsabilité ne se lit pas."""
    journal = _journal([_versement(1, montant="50000", caissier=None)])
    assert journal.lines[0].cashier == "—"


def test_le_recapitulatif_par_caissier_separe_les_deux_caisses() -> None:
    journal = _journal(
        [
            _versement(1, montant="50000", caissier=SOPHIE),
            _versement(2, montant="30000", caissier=SOPHIE),
            _versement(3, montant="120000", caissier=MARIAM),
        ]
    )
    par_caissier = {groupe.key: (groupe.count, groupe.total) for groupe in journal.by_cashier}
    assert par_caissier["Sophie Yao"] == (2, Decimal("80000"))
    assert par_caissier["Mariam Diallo"] == (1, Decimal("120000"))


# ---------------------------------------------------------------------------
# Les totaux
# ---------------------------------------------------------------------------


def test_le_total_est_la_somme_exacte_des_versements_valides() -> None:
    journal = _journal(
        [
            _versement(1, montant="50000"),
            _versement(2, montant="75000"),
        ]
    )
    assert journal.total_encaisse == Decimal("125000")


def test_un_versement_annule_est_compte_mais_jamais_additionne() -> None:
    """Il figure au détail parce qu'il s'est passé quelque chose, et son
    montant n'entre dans aucun total."""
    journal = _journal(
        [
            _versement(1, montant="50000"),
            _versement(2, montant="60000", status="cancelled"),
            _versement(3, montant="30000", status="pending"),
        ]
    )
    assert journal.total_encaisse == Decimal("50000")
    assert len(journal.lines) == 3
    assert journal.counts_by_status == {"completed": 1, "cancelled": 1, "pending": 1}


def test_la_ventilation_par_moyen_redonne_toujours_le_total() -> None:
    journal = _journal(
        [
            _versement(1, montant="50000", method="cash"),
            _versement(2, montant="75000", method="mobile_money"),
            _versement(3, montant="120000", method="bank_transfer"),
        ]
    )
    assert sum(groupe.total for groupe in journal.by_method) == journal.total_encaisse


def test_un_moyen_de_paiement_inconnu_garde_sa_propre_ligne() -> None:
    """Un moyen ajouté après coup ne doit ni disparaître ni tomber dans un
    fourre-tout : c'est ce qui fausse un total sans le dire."""
    journal = _journal(
        [
            _versement(1, montant="50000", method="cash"),
            _versement(2, montant="45000", method="orange_money"),
            _versement(3, montant="55000", method="wave"),
        ]
    )
    par_moyen = {groupe.key: groupe.total for groupe in journal.by_method}
    assert par_moyen["orange_money"] == Decimal("45000")
    assert par_moyen["wave"] == Decimal("55000")
    assert sum(par_moyen.values()) == journal.total_encaisse == Decimal("150000")


def test_un_etat_inconnu_est_compte_sous_son_propre_nom() -> None:
    journal = _journal([_versement(1, montant="50000", status="litigious")])
    assert journal.counts_by_status == {"litigious": 1}
    assert journal.total_encaisse == Decimal("0")


def test_une_selection_tronquee_le_dit() -> None:
    """Un document incomplet qui se tait est pire qu'un document absent : on
    le signe en croyant qu'il est complet."""
    journal = _journal([_versement(1, montant="50000")], total_found=4200)
    assert journal.truncated_from == 4200


# ---------------------------------------------------------------------------
# Le document PDF
# ---------------------------------------------------------------------------


def test_le_document_porte_l_en_tete_de_l_etablissement() -> None:
    journal = _journal([_versement(1, montant="50000")])
    html = render_payments_journal_html(journal, ECOLE)
    assert "Lycée Moderne Saint-Augustin" in html
    assert "Boulevard Latrille, Cocody" in html
    assert "JOURNAL DES VERSEMENTS" in html
    assert "RÉPUBLIQUE DE CÔTE D'IVOIRE" in html.upper()


def test_le_document_prend_les_couleurs_configurees_par_l_ecole() -> None:
    journal = _journal([_versement(1, montant="50000")])
    html = render_payments_journal_html(journal, ECOLE)
    assert "#1B4D3E" in html, "la couleur primaire du tenant doit piloter le document"
    assert "#C1440E" in html


def test_le_document_nomme_qui_a_encaisse_et_dit_ce_qu_il_couvre() -> None:
    journal = _journal(
        [_versement(1, montant="50000", caissier=MARIAM)],
        scope_label="Ma caisse — Mariam Diallo",
        filters_label="État : Validé",
    )
    html = render_payments_journal_html(journal, ECOLE)
    assert "Encaissé par" in html
    assert "Mariam Diallo" in html
    assert "Ma caisse — Mariam Diallo" in html
    assert "État : Validé" in html
    assert "Du 14/09/2026 au 16/09/2026" in html


def test_le_document_montre_le_total_en_francs_lisibles() -> None:
    journal = _journal([_versement(1, montant="1250000")])
    html = render_payments_journal_html(journal, ECOLE)
    # Espaces insécables : un montant ne doit jamais se couper en fin de ligne.
    assert "1 250 000" in html
    assert "1250000" not in html, "un montant collé se relit mal sur un document de caisse"


def test_le_pdf_sort_vraiment_avec_l_en_tete() -> None:
    """Le rendu natif n'est pas disponible partout ; quand il l'est, on vérifie
    que le document se produit réellement."""
    weasyprint = pytest.importorskip("weasyprint")
    from app.services.pdf.payments_journal import generate_payments_journal_pdf

    journal = _journal([_versement(1, montant="50000")])
    try:
        contenu = generate_payments_journal_pdf(journal, ECOLE)
    except OSError as exc:  # bibliothèques natives absentes (GTK/Cairo)
        pytest.skip(f"rendu natif indisponible : {exc}")
    assert contenu.startswith(b"%PDF")
    assert len(contenu) > 1000
    assert weasyprint is not None


# ---------------------------------------------------------------------------
# Le classeur Excel
# ---------------------------------------------------------------------------


def _classeur(journal):
    return load_workbook(io.BytesIO(generate_payments_journal_xlsx(journal, ECOLE)))


def _valeurs(ws) -> list[object]:
    return [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]


def test_le_classeur_porte_l_identite_de_l_ecole() -> None:
    journal = _journal([_versement(1, montant="50000")])
    valeurs = _valeurs(_classeur(journal)["Journal"])
    assert "Lycée Moderne Saint-Augustin" in valeurs
    assert "Journal des versements" in valeurs
    assert any("Édité le" in str(v) for v in valeurs)


def test_le_classeur_prend_la_couleur_de_l_ecole_sur_ses_en_tetes() -> None:
    journal = _journal([_versement(1, montant="50000")])
    ws = _classeur(journal)["Journal"]
    entete = next(
        row for row in ws.iter_rows() if any(cell.value == "Encaissé par" for cell in row)
    )
    couleurs = {cell.fill.start_color.rgb for cell in entete if cell.value}
    assert "FF1B4D3E" in couleurs, "l'en-tête doit porter la couleur primaire du tenant"


def test_le_classeur_nomme_qui_a_encaisse_chaque_ligne() -> None:
    journal = _journal(
        [
            _versement(1, montant="50000", caissier=SOPHIE),
            _versement(2, montant="120000", caissier=MARIAM),
        ]
    )
    valeurs = _valeurs(_classeur(journal)["Journal"])
    assert "Sophie Yao" in valeurs
    assert "Mariam Diallo" in valeurs


def test_les_montants_du_classeur_sont_des_nombres_pas_du_texte() -> None:
    """Un export dont on ne peut pas refaire la somme dans le tableur ne sert
    à rien."""
    journal = _journal([_versement(1, montant="50000"), _versement(2, montant="75000")])
    ws = _classeur(journal)["Journal"]
    montants = [
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, int | float) and cell.value in (50000, 75000, 125000)
    ]
    assert 50000 in montants
    assert 75000 in montants
    assert 125000 in montants, "la ligne de total doit être un nombre elle aussi"


def test_le_recapitulatif_du_classeur_ventile_par_moyen_et_par_caissier() -> None:
    journal = _journal(
        [
            _versement(1, montant="50000", method="cash", caissier=SOPHIE),
            _versement(2, montant="45000", method="orange_money", caissier=MARIAM),
        ]
    )
    valeurs = _valeurs(_classeur(journal)["Récapitulatif"])
    assert "Espèces" in valeurs
    assert "orange_money" in valeurs, "un moyen inconnu reste visible sous son nom"
    assert "Sophie Yao" in valeurs
    assert "Mariam Diallo" in valeurs
