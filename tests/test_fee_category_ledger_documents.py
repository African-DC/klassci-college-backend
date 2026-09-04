"""Le PDF et le classeur sortent du meme point : ils doivent dire la meme chose.

Le comptable les lit cote a cote — il recalcule dans le tableur ce qu'il fait
signer en PDF. Ils partaient pourtant chacun de leur cote : colonnes
differentes, ligne de total d'un seul cote, date d'edition d'un seul cote, et
un tiret qui ne voulait pas dire la meme chose ici et la-bas. Le meme eleve
sortait « inconnu » dans la piece signee et « solde » dans celle qu'on
recalcule.

Ces tests-la ne peuvent pas se remplacer par des tests de chaque sortie prise
seule : ce qu'ils mesurent est un ECART, et un ecart ne se voit qu'a deux.
"""

import io
from decimal import Decimal

from openpyxl import load_workbook

from app.services.exports.fee_category_ledger_xlsx import generate_fee_category_ledger_xlsx
from app.services.payments import ledger_labels
from app.services.pdf._helpers import format_xof
from app.services.pdf.fee_category_ledger import render_fee_category_ledger_html
from tests.fee_category_ledger_decor import CAISSIERE, ECOLE, document, ligne


def _feuille(point) -> object:
    """Le classeur relu comme le comptable l'ouvre : par ses cellules."""
    return load_workbook(io.BytesIO(generate_fee_category_ledger_xlsx(point, ECOLE))).active


def _textes(feuille) -> list[str]:
    """Toutes les cellules textuelles, a plat : ce que le classeur DIT."""
    return [
        str(cellule.value)
        for row in feuille.iter_rows()
        for cellule in row
        if isinstance(cellule.value, str)
    ]


def _entetes(feuille, colonnes: int) -> list[str]:
    """La ligne d'en-tete du tableau, retrouvee par son premier libelle."""
    for row in feuille.iter_rows():
        if row and row[0].value == "Élève":
            return [str(cellule.value) for cellule in row[:colonnes]]
    raise AssertionError("Le classeur ne porte aucun en-tête de tableau.")


# ---------------------------------------------------------------------------
# Les memes colonnes des deux cotes
# ---------------------------------------------------------------------------


def test_les_deux_sorties_portent_les_memes_colonnes() -> None:
    """Le classeur avait une colonne « Déposé le » que le PDF n'avait pas."""
    point = document(consolide=True)
    colonnes = ledger_labels.colonnes(consolide=True, accepts_in_kind=True)

    html = render_fee_category_ledger_html(point, ECOLE)
    entetes = _entetes(_feuille(point), len(colonnes))

    assert entetes == [colonne.label for colonne in colonnes]
    for colonne in colonnes:
        assert colonne.label in html


def test_la_colonne_du_depot_disparait_des_deux_cotes_sans_depot_en_nature() -> None:
    point = document(consolide=True, accepts_in_kind=False)

    assert "Déposé le" not in render_fee_category_ledger_html(point, ECOLE)
    assert "Déposé le" not in _textes(_feuille(point))


def test_la_colonne_du_reste_disparait_des_deux_cotes_pour_une_caissiere() -> None:
    """Elle sortait integralement remplie de tirets sous un en-tete en francs.

    Une colonne qu'on ne peut pas remplir se retire ; elle ne se remplit pas de
    tirets. C'est la meme regle que `remaining=None` dans le contrat HTTP :
    absent vaut mieux que faux.
    """
    point = document(consolide=False)

    assert "Reste à payer (XOF)" not in render_fee_category_ledger_html(point, ECOLE)
    assert "Reste à payer (XOF)" not in _textes(_feuille(point))


# ---------------------------------------------------------------------------
# Les memes mots
# ---------------------------------------------------------------------------


def test_les_deux_sorties_ecrivent_le_meme_mot_pour_un_etat() -> None:
    """Deux tables d'etats privees finissent toujours par diverger."""
    point = document(consolide=True, lignes=(ligne(status="in_kind"),))

    assert ledger_labels.etat_label("in_kind") in render_fee_category_ledger_html(point, ECOLE)
    assert ledger_labels.etat_label("in_kind") in _textes(_feuille(point))


def test_les_deux_sorties_avertissent_du_cloisonnement_dans_les_memes_termes() -> None:
    """Le PDF et le classeur disaient la meme chose en deux phrases differentes."""
    point = document(consolide=False)

    assert ledger_labels.AVERTISSEMENT_CAISSE in render_fee_category_ledger_html(point, ECOLE)
    assert ledger_labels.AVERTISSEMENT_CAISSE in _textes(_feuille(point))


def test_les_deux_sorties_repetent_le_meme_rappel_en_pied() -> None:
    """« n'y figurent pas » d'un cote, « ne peuvent pas y figurer » de l'autre."""
    point = document(consolide=False)

    assert ledger_labels.RAPPEL_CAISSE in render_fee_category_ledger_html(point, ECOLE)
    assert ledger_labels.RAPPEL_CAISSE in _textes(_feuille(point))


def test_les_deux_sorties_disent_quels_versements_elles_comptent() -> None:
    """Un total qui ne le dit pas laisse croire qu'il compte les versements saisis."""
    point = document(consolide=True)

    assert ledger_labels.VERSEMENTS_COMPTES in render_fee_category_ledger_html(point, ECOLE)
    assert any(ledger_labels.VERSEMENTS_COMPTES in texte for texte in _textes(_feuille(point)))


# ---------------------------------------------------------------------------
# La meme identite, et la meme date d'edition
# ---------------------------------------------------------------------------


def test_les_deux_sorties_nomment_l_annee_le_perimetre_et_la_caisse() -> None:
    point = document(consolide=False)
    html = render_fee_category_ledger_html(point, ECOLE)
    textes = _textes(_feuille(point))

    for attendu in ("2026-2027", "Toutes les classes", f"Ma caisse — {CAISSIERE}"):
        assert attendu in html
        assert any(attendu in texte for texte in textes)


def test_le_classeur_porte_lui_aussi_sa_date_d_edition() -> None:
    """Il ne la portait pas du tout : deux tirages a deux jours etaient identiques."""
    point = document(consolide=True)
    attendu = ledger_labels.issued_label(point.issued_at, point.issued_by)

    assert attendu in render_fee_category_ledger_html(point, ECOLE)
    assert attendu in _textes(_feuille(point))


# ---------------------------------------------------------------------------
# Le meme total, et la meme mention de troncature
# ---------------------------------------------------------------------------


def test_le_pdf_porte_lui_aussi_la_ligne_de_total() -> None:
    """Elle n'existait que dans le classeur, alors que le PDF est la piece signee."""
    point = document(consolide=True)

    assert ledger_labels.TOTAL_LABEL in render_fee_category_ledger_html(point, ECOLE)
    assert ledger_labels.TOTAL_LABEL in _textes(_feuille(point))


def test_la_ligne_de_total_porte_le_total_du_perimetre_des_deux_cotes() -> None:
    point = document(consolide=True)
    feuille = _feuille(point)

    montants = [
        cellule.value
        for row in feuille.iter_rows()
        for cellule in row
        if isinstance(cellule.value, (int, float))
    ]
    assert float(point.total_en_argent) in montants
    # Le PDF affiche le meme chiffre, mis en forme par le formateur maison —
    # dont le separateur de milliers est insecable, ce qu'un litteral tape a
    # la main dans ce test ne serait pas.
    assert format_xof(point.total_en_argent) in render_fee_category_ledger_html(point, ECOLE)


def test_un_point_sans_ligne_ne_porte_pas_de_total_et_le_dit() -> None:
    """Une ligne de total sous zero ligne se lit comme un total de zero encaisse."""
    point = document(consolide=True, lignes=())
    html = render_fee_category_ledger_html(point, ECOLE)
    textes = _textes(_feuille(point))

    assert ledger_labels.AUCUNE_LIGNE in html
    assert ledger_labels.AUCUNE_LIGNE in textes
    assert ledger_labels.TOTAL_LABEL not in html
    assert ledger_labels.TOTAL_LABEL not in textes


def test_les_deux_sorties_annoncent_la_troncature() -> None:
    """Le plafond coupe la liste ; un document ampute qui se tait se signe quand meme."""
    point = document(consolide=True, truncated_from=6200)
    attendu = ledger_labels.troncature_label(len(point.lignes), 6200)

    assert attendu in render_fee_category_ledger_html(point, ECOLE)
    assert attendu in _textes(_feuille(point))


# ---------------------------------------------------------------------------
# Le tiret dit « on ne sait pas », des deux cotes
# ---------------------------------------------------------------------------


def test_un_montant_nul_s_ecrit_zero_et_jamais_un_tiret() -> None:
    """Le PDF ecrivait « — » sous une ligne soldee, le classeur « 0 F »."""
    point = document(
        consolide=True,
        lignes=(ligne(status="paid", paid=Decimal("3000"), remaining=Decimal("0")),),
    )
    feuille = _feuille(point)

    # Aucune cellule de montant ne porte le tiret : les montants sont des
    # nombres, y compris le zero.
    entetes = _entetes(feuille, 8)
    colonne_reste = entetes.index("Reste à payer (XOF)") + 1
    valeurs = [
        feuille.cell(row=row, column=colonne_reste).value for row in range(1, feuille.max_row + 1)
    ]
    assert ledger_labels.ABSENT not in valeurs
    assert 0.0 in valeurs

    assert '<td class="num">0</td>' in render_fee_category_ledger_html(point, ECOLE)


def test_un_depot_absent_reste_un_tiret_des_deux_cotes() -> None:
    """La, le tiret est juste : la ligne n'a pas ete deposee, on ne devine pas."""
    point = document(consolide=True)

    assert ledger_labels.ABSENT in render_fee_category_ledger_html(point, ECOLE)
    assert ledger_labels.ABSENT in _textes(_feuille(point))
