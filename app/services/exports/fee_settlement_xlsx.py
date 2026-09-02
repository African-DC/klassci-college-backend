"""Le tableau des soldes au format classeur.

Une seule feuille : le tableau est déjà une grille, et le récapitulatif tient
en une ligne d'en-tête. Ce document se relit hors ligne, au bureau, sur la
classe entière — c'est pour ça qu'il existe plutôt qu'un simple écran.

Le reste dû est écrit en nombre à côté de l'état, jamais fondu dedans : « Dû »
ne dit pas combien, et un fondateur qui prépare ses relances a besoin des deux.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.exports._workbook_branding import (
    set_widths,
    setup_printing,
    style_table_header,
    write_header,
)
from app.services.fee_settlement import (
    STATE_LABEL,
    SettlementMatrix,
    SettlementState,
)

_MONEY_FORMAT = '#,##0" F"'
#: Largeur des deux colonnes d'identité, puis de chaque colonne de catégorie.
_WIDTH_NAME = 28
_WIDTH_MATRICULE = 16
_WIDTH_CATEGORY = 18
_WIDTH_TOTAL = 18


def _meta_lines(matrix: SettlementMatrix) -> list[str]:
    return [
        f"{matrix.settled_count} élèves soldés sur {matrix.total_count}",
        "Un élève est soldé quand plus rien n'est dû en argent, "
        "dépôts en nature et exonérations compris.",
    ]


def _write_matrix(ws: Worksheet, matrix: SettlementMatrix, school: dict[str, Any]) -> None:
    headers = ["Élève", "Matricule"]
    headers += [colonne.name for colonne in matrix.columns]
    headers += ["Reste à payer (XOF)", "Soldé"]

    header_row = write_header(
        ws,
        school,
        title="Soldes par catégorie de frais",
        subtitle=f"{matrix.class_name} — {matrix.academic_year_name}".strip(" —"),
        meta_lines=_meta_lines(matrix),
        width=len(headers),
    )
    set_widths(
        ws,
        [_WIDTH_NAME, _WIDTH_MATRICULE]
        + [_WIDTH_CATEGORY] * len(matrix.columns)
        + [_WIDTH_TOTAL, 10],
    )

    for index, label in enumerate(headers, 1):
        ws.cell(row=header_row, column=index, value=label)
    style_table_header(ws, header_row, len(headers), school)

    row = header_row
    for ligne in matrix.rows:
        row += 1
        ws.cell(row=row, column=1, value=f"{ligne.last_name} {ligne.first_name}".strip())
        ws.cell(row=row, column=2, value=ligne.student_matricule or "—")
        for offset, cell in enumerate(ligne.cells, 3):
            # L'état sous son propre nom, et le reste dû à côté quand il y en
            # a. « Partiel » seul obligerait à rouvrir la fiche pour savoir
            # combien réclamer.
            label = STATE_LABEL.get(cell.state, cell.state.value)
            if cell.state == SettlementState.PARTIAL:
                label = f"{label} — reste {cell.remaining:,.0f} F".replace(",", " ")
            ws.cell(row=row, column=offset, value=label)
        reste = ws.cell(
            row=row,
            column=2 + len(matrix.columns) + 1,
            value=float(sum(cell.remaining for cell in ligne.cells)),
        )
        reste.number_format = _MONEY_FORMAT
        ws.cell(
            row=row,
            column=2 + len(matrix.columns) + 2,
            value="Oui" if ligne.settled else "Non",
        )

    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)
    setup_printing(ws, header_row=header_row)


def generate_fee_settlement_xlsx(
    matrix: SettlementMatrix, school_settings: dict[str, Any]
) -> bytes:
    """Génère le classeur des soldes aux couleurs de l'école."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Soldes"
    _write_matrix(sheet, matrix, school_settings)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
