"""Identité de l'établissement dans un classeur Excel.

Un export comptable circule : il est envoyé au fondateur, imprimé, joint à un
courrier. Sans nom d'école, sans logo, sans couleurs, il ressemble à un
brouillon exporté à la va-vite — et surtout, détaché de son contexte, on ne
sait plus de quel établissement ni de quelle période il parle.

Les couleurs sortent des paramètres du tenant, comme pour les PDF. Le repli sur
le bleu KLASSCI n'existe que pour l'école qui n'a rien configuré.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.pdf._helpers import image_bytes
from app.services.pdf.theme import DEFAULT_ACCENT, DEFAULT_PRIMARY

logger = logging.getLogger(__name__)

_HEX_RE = re.compile(r"^#?([0-9A-Fa-f]{6})$")

#: Première ligne occupée par la bande d'identité.
_FIRST_META_ROW = 5

THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def argb(color: str | None, fallback: str) -> str:
    """Normalise une couleur de paramétrage en ARGB accepté par openpyxl.

    Une valeur invalide n'est pas une raison de faire échouer un export : on
    retombe sur la couleur par défaut, le document sort, et il reste lisible.
    """
    match = _HEX_RE.match((color or "").strip())
    if match is None:
        match = _HEX_RE.match(fallback)
    return f"FF{match.group(1).upper()}" if match else "FF0F3F8C"


def primary_argb(school: dict[str, Any]) -> str:
    return argb(school.get("primary_color"), DEFAULT_PRIMARY)


def accent_argb(school: dict[str, Any]) -> str:
    return argb(school.get("accent_color"), DEFAULT_ACCENT)


def write_header(
    ws: Worksheet,
    school: dict[str, Any],
    *,
    title: str,
    subtitle: str,
    meta_lines: list[str],
    width: int,
) -> int:
    """Écrit la bande d'identité et renvoie la première ligne libre en dessous.

    Renvoyer la ligne plutôt que de la faire recalculer par l'appelant évite
    la classe de bug la plus bête d'un export : un en-tête qui grandit d'une
    ligne et un tableau qui vient écrire par-dessus.

    `width` est le nombre de colonnes du tableau qui suit, pour que les
    fusions de l'en-tête couvrent exactement la largeur utile.
    """
    primary = primary_argb(school)
    last_col = get_column_letter(max(width, 3))
    # La première colonne n'est réservée au logo que s'il y en a un. Sinon
    # l'identité de l'école commencerait par une colonne vide, et un lecteur
    # y verrait un décalage plutôt qu'une place gardée.
    first_col = "B" if _insert_logo(ws, school) else "A"

    lignes = [
        (school.get("school_name") or "Établissement", Font(size=15, bold=True, color=primary)),
        (
            " · ".join(
                str(school[key]) for key in ("address", "phone", "email") if school.get(key)
            ),
            Font(size=9, color="FF666666"),
        ),
        (title, Font(size=12, bold=True)),
        (subtitle, Font(size=10, italic=True, color="FF555555")),
    ]
    for row, (value, font) in enumerate(lignes, 1):
        ws.merge_cells(f"{first_col}{row}:{last_col}{row}")
        cell = ws[f"{first_col}{row}"]
        cell.value = value
        cell.font = font
        cell.alignment = LEFT

    row = _FIRST_META_ROW
    for line in meta_lines:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        ws[f"A{row}"].value = line
        ws[f"A{row}"].font = Font(size=9, color="FF555555")
        ws[f"A{row}"].alignment = LEFT
        row += 1

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 18
    # Une ligne vide entre l'identité et le tableau : sans elle, l'en-tête
    # colorée du tableau se colle au bloc école et les deux se confondent.
    return row + 1


def _insert_logo(ws: Worksheet, school: dict[str, Any]) -> bool:
    """Pose le logo en A1 et dit s'il a réellement été posé.

    Son absence ne compromet jamais l'export : le classeur reste identifié par
    le nom et les couleurs de l'école.
    """
    resolved = image_bytes(school.get("logo_url"))
    if resolved is None:
        return False
    content, _mime = resolved
    try:
        from openpyxl.drawing.image import Image as XlsxImage

        image = XlsxImage(io.BytesIO(content))
        ratio = image.height / image.width if image.width else 1
        image.width = 90
        image.height = max(int(90 * ratio), 1)
        ws.column_dimensions["A"].width = 14
        ws.add_image(image, "A1")
        return True
    except Exception:  # pragma: no cover — format d'image exotique
        logger.warning("Logo non inséré dans l'export Excel", exc_info=True)
        return False


def setup_printing(ws: Worksheet, *, header_row: int, landscape: bool = True) -> None:
    """Rend la feuille imprimable d'un bloc, en-têtes répétées.

    Un classeur comptable finit imprimé. Sans ces réglages, le tableau se
    coupe en tranches de colonnes réparties sur plusieurs feuilles, et les
    pages suivantes arrivent sans titres de colonnes : des montants sans
    étiquette, donc inexploitables.
    """
    if landscape:
        ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"


def style_table_header(ws: Worksheet, row: int, width: int, school: dict[str, Any]) -> None:
    """Applique la couleur primaire de l'école à une ligne d'en-têtes."""
    fill = PatternFill(
        start_color=primary_argb(school), end_color=primary_argb(school), fill_type="solid"
    )
    for col in range(1, width + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = Font(color="FFFFFFFF", bold=True, size=10)
        cell.alignment = CENTER
    ws.row_dimensions[row].height = 20


def style_total_row(ws: Worksheet, row: int, width: int, school: dict[str, Any]) -> None:
    """Met en évidence une ligne de total avec la couleur d'accent de l'école."""
    accent = accent_argb(school)
    fill = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
    for col in range(1, width + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFFFF", size=11)
    ws.row_dimensions[row].height = 20


def set_widths(ws: Worksheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
