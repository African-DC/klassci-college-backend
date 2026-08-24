"""Export Excel des statistiques DREN (openpyxl).

Produit un classeur à trois feuilles : Synthèse (indicateurs clés), Niveaux &
classes (effectifs détaillés), Matières (moyennes générales). Utilisé par
l'endpoint d'export DREN au format `xlsx`.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill(start_color="0F3F8C", end_color="0F3F8C", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=14, bold=True, color="0F3F8C")
_TOTAL_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center")


def _num(value: Any) -> float | str:
    return float(value) if value is not None else "—"


def _pct(value: float | None) -> str:
    return f"{value:.1f} %" if value is not None else "—"


def _style_header(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def generate_dren_stats_xlsx(data: Any) -> bytes:
    """Génère le classeur Excel des statistiques DREN.

    `data` : instance de `DrenStatsResponse` (accès par attributs).
    """
    wb = Workbook()

    # --- Feuille 1 : Synthèse -------------------------------------------------
    ws = wb.active
    ws.title = "Synthèse"
    ws["A1"] = "Statistiques DREN"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Année scolaire {data.academic_year_name}"
    ws["A2"].font = Font(italic=True, color="555555")

    rows = [
        ("Effectif total", data.total_students),
        ("Garçons", data.male_count),
        ("Filles", data.female_count),
        ("Taux de réussite", _pct(data.success_rate)),
        ("Taux d'échec", _pct(data.failure_rate)),
        ("Taux de redoublement", _pct(data.redoublement_rate)),
        ("Taux d'exclusion", _pct(data.exclusion_rate)),
    ]
    start = 4
    ws.cell(row=start, column=1, value="Indicateur")
    ws.cell(row=start, column=2, value="Valeur")
    _style_header(ws, start, 2)
    for i, (label, value) in enumerate(rows, start + 1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
    _autosize(ws, [26, 18])

    # --- Feuille 2 : Niveaux & classes ---------------------------------------
    ws2 = wb.create_sheet("Niveaux & classes")
    headers = ["Niveau", "Classe", "Effectif", "Garçons", "Filles", "Moyenne /20"]
    ws2.append(headers)
    _style_header(ws2, 1, len(headers))
    for lvl in data.levels:
        if not lvl.classes:
            ws2.append(
                [lvl.level_name, "—", lvl.total_students, lvl.male_count, lvl.female_count, "—"]
            )
            continue
        for cls in lvl.classes:
            ws2.append(
                [
                    lvl.level_name,
                    cls.class_name,
                    cls.total_students,
                    cls.male_count,
                    cls.female_count,
                    _num(cls.average),
                ]
            )
    total_row = [
        "Total établissement",
        "",
        data.total_students,
        data.male_count,
        data.female_count,
        "",
    ]
    ws2.append(total_row)
    for col in range(1, len(headers) + 1):
        ws2.cell(row=ws2.max_row, column=col).font = _TOTAL_FONT
    _autosize(ws2, [18, 20, 12, 12, 12, 14])

    # --- Feuille 3 : Matières -------------------------------------------------
    ws3 = wb.create_sheet("Matières")
    headers3 = ["Matière", "Moyenne générale /20", "Enseignants"]
    ws3.append(headers3)
    _style_header(ws3, 1, len(headers3))
    for subj in data.subjects:
        ws3.append([subj.subject_name, _num(subj.overall_average), subj.teacher_count])
    _autosize(ws3, [28, 22, 14])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
