"""Journal des versements au format classeur Excel.

Deux feuilles, parce que deux lectures cohabitent : le détail, ligne à ligne,
qu'on relit pour retrouver un versement ; et le récapitulatif, qu'on lit pour
contrôler une caisse. Les montants sont écrits en nombres, pas en texte — un
export dont on ne peut pas refaire la somme dans le tableur ne sert à rien.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from app.services.exports._workbook_branding import (
    set_widths,
    setup_printing,
    style_table_header,
    style_total_row,
    write_header,
)
from app.services.payments.journal_data import PaymentsJournal
from app.services.payments.journal_labels import fee_cell
from app.services.pdf.theme import method_label, status_label

# Le séparateur de milliers est rendu par le tableur selon la locale du
# lecteur : espace en français, virgule ailleurs. L'écrire en dur donnerait un
# classeur juste sur le poste qui l'a produit et faux sur celui qui l'ouvre.
_MONEY_FORMAT = '#,##0" F"'

_DETAIL_HEADERS = [
    "N°",
    "Date",
    "Heure",
    "Élève",
    "Matricule",
    "Frais",
    "Moyen de paiement",
    "Référence",
    "Encaissé par",
    "État",
    "Montant (XOF)",
]
_DETAIL_WIDTHS = [8, 12, 8, 28, 16, 30, 20, 18, 24, 14, 16]


def _meta_lines(journal: PaymentsJournal) -> list[str]:
    lignes = [f"Périmètre : {journal.scope_label}"]
    if journal.filters_label:
        lignes.append(f"Filtres appliqués : {journal.filters_label}")
    lignes.append(f"Édité le {journal.issued_at.strftime('%d/%m/%Y à %H:%M')}")
    if journal.truncated_from is not None:
        lignes.append(
            f"Attention : {len(journal.lines)} versements présentés sur "
            f"{journal.truncated_from} retenus. Resserrez la période."
        )
    return lignes


def _write_detail(ws: Worksheet, journal: PaymentsJournal, school: dict[str, Any]) -> None:
    header_row = write_header(
        ws,
        school,
        title="Journal des versements",
        subtitle=journal.period_label,
        meta_lines=_meta_lines(journal),
        width=len(_DETAIL_HEADERS),
    )
    set_widths(ws, _DETAIL_WIDTHS)

    for index, label in enumerate(_DETAIL_HEADERS, 1):
        ws.cell(row=header_row, column=index, value=label)
    style_table_header(ws, header_row, len(_DETAIL_HEADERS), school)

    row = header_row
    for line in journal.lines:
        row += 1
        ws.cell(row=row, column=1, value=line.id)
        ws.cell(row=row, column=2, value=line.created_at.strftime("%d/%m/%Y"))
        ws.cell(row=row, column=3, value=line.created_at.strftime("%H:%M"))
        ws.cell(row=row, column=4, value=line.student_name)
        ws.cell(row=row, column=5, value=line.student_matricule or "—")
        frais = ws.cell(row=row, column=6, value=fee_cell(line.fee_shares))
        # Sans le renvoi à la ligne, les trois catégories d'un versement
        # réparti tiennent sur une seule ligne écrasée : le tableur ne coupe
        # pas de lui-même sur un retour à la ligne.
        frais.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=7, value=method_label(line.method))
        ws.cell(row=row, column=8, value=line.reference or "—")
        ws.cell(row=row, column=9, value=line.cashier)
        ws.cell(row=row, column=10, value=status_label(line.status))
        montant = ws.cell(row=row, column=11, value=float(line.amount))
        montant.number_format = _MONEY_FORMAT

    total_row = row + 1
    ws.cell(row=total_row, column=1, value="Total encaissé (versements validés)")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=10)
    total_cell = ws.cell(row=total_row, column=11, value=float(journal.total_encaisse))
    total_cell.number_format = _MONEY_FORMAT
    style_total_row(ws, total_row, len(_DETAIL_HEADERS), school)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    setup_printing(ws, header_row=header_row)


def _write_group(
    ws: Worksheet,
    *,
    start_row: int,
    title: str,
    key_header: str,
    groups: list[Any],
    labeller: Any,
    school: dict[str, Any],
) -> int:
    """Écrit un bloc récapitulatif et renvoie la ligne suivante libre."""
    ws.cell(row=start_row, column=1, value=title).font = Font(size=11, bold=True)
    header_row = start_row + 1
    for index, label in enumerate([key_header, "Versements", "Total (XOF)"], 1):
        ws.cell(row=header_row, column=index, value=label)
    style_table_header(ws, header_row, 3, school)

    row = header_row
    total = 0.0
    for group in groups:
        row += 1
        ws.cell(row=row, column=1, value=labeller(group.key))
        ws.cell(row=row, column=2, value=group.count)
        montant = ws.cell(row=row, column=3, value=float(group.total))
        montant.number_format = _MONEY_FORMAT
        total += float(group.total)

    row += 1
    ws.cell(row=row, column=1, value="Total")
    ws.cell(row=row, column=2, value=sum(group.count for group in groups))
    total_cell = ws.cell(row=row, column=3, value=total)
    total_cell.number_format = _MONEY_FORMAT
    style_total_row(ws, row, 3, school)
    return row + 3


def _write_summary(ws: Worksheet, journal: PaymentsJournal, school: dict[str, Any]) -> None:
    row = write_header(
        ws,
        school,
        title="Récapitulatif des versements",
        subtitle=journal.period_label,
        meta_lines=_meta_lines(journal),
        width=3,
    )
    set_widths(ws, [34, 16, 20])

    row = _write_group(
        ws,
        start_row=row,
        title="Par moyen de paiement",
        key_header="Moyen de paiement",
        groups=journal.by_method,
        labeller=method_label,
        school=school,
    )
    row = _write_group(
        ws,
        start_row=row,
        title="Par caissier",
        key_header="Encaissé par",
        groups=journal.by_cashier,
        labeller=str,
        school=school,
    )

    ws.cell(row=row, column=1, value="Décompte par état").font = Font(size=11, bold=True)
    header_row = row + 1
    for index, label in enumerate(["État", "Versements"], 1):
        ws.cell(row=header_row, column=index, value=label)
    style_table_header(ws, header_row, 2, school)
    for offset, (status, count) in enumerate(sorted(journal.counts_by_status.items()), 1):
        ws.cell(row=header_row + offset, column=1, value=status_label(status))
        ws.cell(row=header_row + offset, column=2, value=count)

    setup_printing(ws, header_row=1, landscape=False)


def generate_payments_journal_xlsx(
    journal: PaymentsJournal, school_settings: dict[str, Any]
) -> bytes:
    """Génère le classeur du journal des versements aux couleurs de l'école."""
    wb = Workbook()
    detail = wb.active
    detail.title = "Journal"
    _write_detail(detail, journal, school_settings)

    summary = wb.create_sheet("Récapitulatif")
    _write_summary(summary, journal, school_settings)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
