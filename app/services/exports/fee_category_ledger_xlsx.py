"""Le point d'une catégorie de frais, au format classeur.

C'est le document que le comptable « tire ». Il sort de l'école : il part chez
un prestataire pour justifier un versement, ou sur un bureau pour préparer des
relances. Il doit donc dire de quoi il parle sans qu'on ait à le demander — la
catégorie, l'année, la période, le périmètre, la caisse, qui l'a tiré et quand
— et surtout **dire ce qu'il ne dit pas**.

Un classeur qui tairait qu'il ne couvre qu'une seule caisse serait pris pour le
compte de l'école entière. C'est la première ligne de son en-tête.

**Ses mots ne sont pas les siens** : ils viennent de
`app.services.payments.ledger_labels`, partagés avec le PDF. Les deux sortent
du même point et se lisent côte à côte — le comptable recalcule ici ce qu'il
fait signer là-bas — et ils portaient chacun leur table d'états, leurs colonnes
et leurs phrases. Un mot qui diverge d'un fichier à l'autre, sur une pièce
comptable, est un mot faux dans l'un des deux.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.services.exports._workbook_branding import (
    set_widths,
    setup_printing,
    style_table_header,
    style_total_row,
    write_header,
)
from app.services.fee_category_ledger import CategoryLedger
from app.services.payments import ledger_labels as mots

_MONEY = '#,##0" F"'

#: La largeur de chaque colonne, quand elle est là. Les colonnes présentes
#: dépendent des droits du lecteur et de la catégorie — voir
#: `ledger_labels.colonnes`.
_LARGEURS: dict[str, int] = {
    "eleve": 28,
    "matricule": 16,
    "classe": 14,
    "etat": 18,
    "du": 14,
    "entre": 24,
    "reste": 18,
    "depose": 14,
}


def _meta(ledger: CategoryLedger) -> list[str]:
    """L'en-tête : ce que le document couvre, et ce qu'il ne couvre pas.

    L'année, le périmètre et la caisse sont lus du CRITÈRE, jamais reconstitués
    depuis les lignes rendues. La date d'édition manquait tout à fait ici alors
    que le PDF la portait : deux tirages du même écran à deux jours d'écart
    étaient indiscernables une fois imprimés.
    """
    lignes = [
        f"Année scolaire : {ledger.academic_year_name}",
        f"Périmètre : {ledger.class_name}",
        f"Caisse : {ledger.scope_label}",
    ]
    filtres = mots.filters_label(state=ledger.etat_filtre, q=ledger.recherche)
    if filtres:
        lignes.append(f"Filtres appliqués : {filtres}")
    lignes.append(mots.issued_label(ledger.issued_at, ledger.issued_by))

    if not ledger.consolide:
        # La ligne la plus importante de l'en-tete : sans elle, un document de
        # guichet se lit comme le compte de l'ecole entiere.
        lignes.append(mots.AVERTISSEMENT_CAISSE)
    if ledger.truncated_from is not None:
        lignes.append(mots.troncature_label(len(ledger.lignes), ledger.truncated_from))

    lignes.append(mots.entre_meta(ledger.eleves_en_argent, ledger.total_en_argent))
    if ledger.accepts_in_kind:
        lignes.append(mots.depots_label(ledger.depots_en_nature))
    if ledger.total_restant_du is not None and ledger.eleves_restant_du is not None:
        lignes.append(mots.reste_du_label(ledger.eleves_restant_du, ledger.total_restant_du))
    return lignes


def _valeur(ligne: Any, cle: str) -> Any:
    """La valeur d'une colonne, pour une ligne d'élève.

    Les montants sortent en NOMBRES, jamais en texte : un export dont on ne
    peut pas refaire la somme dans le tableur ne sert à rien. Ils ne sont
    jamais absents non plus — la colonne du reste dû n'existe pas quand
    personne ne peut la remplir, plutôt que d'être remplie de tirets sous un
    en-tête qui promet des francs.
    """
    if cle == "eleve":
        return f"{ligne.last_name} {ligne.first_name}".strip()
    if cle == "matricule":
        return ligne.student_matricule or mots.ABSENT
    if cle == "classe":
        return ligne.class_name or mots.ABSENT
    if cle == "etat":
        return mots.etat_label(ligne.status)
    if cle == "du":
        return float(ligne.due)
    if cle == "entre":
        return float(ligne.paid)
    if cle == "reste":
        return float(ligne.remaining)
    return ligne.deposited_at.strftime("%d/%m/%Y") if ligne.deposited_at else mots.ABSENT


def _ecrire_lignes(
    ws: Worksheet,
    ledger: CategoryLedger,
    colonnes: tuple[mots.Colonne, ...],
    *,
    header_row: int,
) -> int:
    """Écrit le détail et renvoie la dernière ligne occupée."""
    row = header_row
    for ligne in ledger.lignes:
        row += 1
        for index, colonne in enumerate(colonnes, 1):
            cellule = ws.cell(row=row, column=index, value=_valeur(ligne, colonne.key))
            if colonne.money:
                cellule.number_format = _MONEY
    return row


def _ecrire_total(
    ws: Worksheet,
    ledger: CategoryLedger,
    colonnes: tuple[mots.Colonne, ...],
    *,
    row: int,
    school: dict[str, Any],
) -> None:
    """La ligne de total du PÉRIMÈTRE, la même que celle du PDF.

    Fusionnée jusqu'à la colonne qui précède le total : une étiquette écrasée
    dans une cellule de nom d'élève ne se lit pas à l'impression.
    """
    cles = [colonne.key for colonne in colonnes]
    colonne_entre = cles.index("entre") + 1
    ws.cell(row=row, column=1, value=mots.TOTAL_LABEL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(colonne_entre - 1, 1))
    total = ws.cell(row=row, column=colonne_entre, value=float(ledger.total_en_argent))
    total.number_format = _MONEY
    if "reste" in cles and ledger.total_restant_du is not None:
        reste = ws.cell(
            row=row, column=cles.index("reste") + 1, value=float(ledger.total_restant_du)
        )
        reste.number_format = _MONEY
    style_total_row(ws, row, len(colonnes), school)


def _ecrire(ws: Worksheet, ledger: CategoryLedger, school: dict[str, Any]) -> None:
    colonnes = mots.colonnes(
        consolide=ledger.consolide,
        accepts_in_kind=ledger.accepts_in_kind,
    )
    header_row = write_header(
        ws,
        school,
        title=f"Point sur : {ledger.category_name}",
        subtitle=mots.period_label(ledger.date_from, ledger.date_to),
        meta_lines=_meta(ledger),
        width=len(colonnes),
    )
    set_widths(ws, [_LARGEURS[colonne.key] for colonne in colonnes])

    for index, colonne in enumerate(colonnes, 1):
        ws.cell(row=header_row, column=index, value=colonne.label)
    style_table_header(ws, header_row, len(colonnes), school)

    row = _ecrire_lignes(ws, ledger, colonnes, header_row=header_row)

    if ledger.lignes:
        row += 1
        _ecrire_total(ws, ledger, colonnes, row=row, school=school)
    else:
        # Un tableau sans ligne et sans phrase se lit comme un export raté. Le
        # PDF dit depuis toujours pourquoi il est vide ; le classeur ne disait
        # rien du tout, et posait une ligne de total sous zéro ligne.
        row += 1
        ws.cell(row=row, column=1, value=mots.AUCUNE_LIGNE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(colonnes))

    if not ledger.consolide:
        # Repete en pied de tableau : un lecteur qui arrive par la fin doit le
        # voir aussi.
        note = ws.cell(row=row + 2, column=1, value=mots.RAPPEL_CAISSE)
        note.font = Font(italic=True, size=9)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    setup_printing(ws, header_row=header_row)


def generate_fee_category_ledger_xlsx(
    ledger: CategoryLedger, school_settings: dict[str, Any]
) -> bytes:
    """Génère le classeur du point de catégorie aux couleurs de l'école."""
    wb = Workbook()
    feuille = wb.active
    feuille.title = "Point"
    _ecrire(feuille, ledger, school_settings)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
